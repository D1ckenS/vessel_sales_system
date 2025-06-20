import io
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import logging

# Set up logging
logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self, title="Report"):
        """Initialize Excel exporter with proper error handling"""
        try:
            self.workbook = openpyxl.Workbook()
            # Remove default sheet and create a new one with proper title
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)
            
            # Create new sheet with clean title (Excel sheet names max 31 chars)
            clean_title = "".join(c for c in title if c.isalnum() or c in (' ', '-', '_'))[:31]
            self.worksheet = self.workbook.create_sheet(title=clean_title or "Report")
            
            self.title = title
            self.current_row = 1
            self.header_count = 0  # Track number of headers for merging
            
            # Set up styles
            self._setup_styles()
            
        except Exception as e:
            logger.error(f"ExcelExporter initialization error: {e}")
            raise Exception(f"Failed to initialize Excel exporter: {str(e)}")
    
    def _setup_styles(self):
        """Set up reusable styles"""
        # Title style
        self.title_font = Font(size=16, bold=True, color="1F4E79")
        self.title_alignment = Alignment(horizontal='center', vertical='center')
        self.title_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        
        # Header style
        self.header_font = Font(size=11, bold=True, color="FFFFFF")
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Border style
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

    def add_title(self, title, subtitle=None):
        """Add title with proper column merging based on number of headers"""
        try:
            # Store title and subtitle for later merging when headers are added
            self.report_title = title
            self.report_subtitle = subtitle
            
            # Add title row
            self.worksheet[f'A{self.current_row}'] = title
            self.worksheet[f'A{self.current_row}'].font = self.title_font
            self.worksheet[f'A{self.current_row}'].alignment = self.title_alignment
            self.worksheet[f'A{self.current_row}'].fill = self.title_fill
            
            self.title_row = self.current_row
            self.current_row += 1
            
            # Add subtitle if provided
            if subtitle:
                self.worksheet[f'A{self.current_row}'] = subtitle
                self.worksheet[f'A{self.current_row}'].font = Font(size=12, italic=True, color="666666")
                self.worksheet[f'A{self.current_row}'].alignment = self.title_alignment
                
                self.subtitle_row = self.current_row
                self.current_row += 1
            else:
                self.subtitle_row = None
            
            # Add spacing
            self.current_row += 1
            
        except Exception as e:
            logger.error(f"Error adding Excel title: {e}")
            pass

    def add_metadata(self, metadata):
        """Add metadata section"""
        try:
            if not metadata:
                return
                
            # Add metadata title
            self.worksheet[f'A{self.current_row}'] = "Report Information"
            self.worksheet[f'A{self.current_row}'].font = Font(size=12, bold=True, color="2C3E50")
            self.current_row += 1
            
            # Add metadata key-value pairs
            for key, value in metadata.items():
                if key and value is not None:
                    # Key in column A
                    key_cell = self.worksheet[f'A{self.current_row}']
                    key_cell.value = f"{key}:"
                    key_cell.font = Font(bold=True)
                    key_cell.border = self.border
                    
                    # Value in column B
                    value_cell = self.worksheet[f'B{self.current_row}']
                    # Handle numeric values properly
                    if isinstance(value, (int, float)):
                        value_cell.value = float(value)
                        value_cell.number_format = '#,##0.000'
                    elif isinstance(value, str) and value.startswith('(') and value.endswith(')'):
                        # Handle negative values in brackets
                        try:
                            numeric_value = -float(value.strip('()').replace(',', ''))
                            value_cell.value = numeric_value
                            value_cell.number_format = '_-[Red](#,##0.000_);#,##0.000'  # Accounting format
                        except ValueError:
                            value_cell.value = str(value)
                    else:
                        value_cell.value = str(value)
                    value_cell.border = self.border
                    
                    self.current_row += 1
            
            # Add spacing
            self.current_row += 1
            
        except Exception as e:
            logger.error(f"Error adding Excel metadata: {e}")
            pass

    def add_headers(self, headers):
        """Add headers and merge title rows to match header count"""
        try:
            if not headers:
                return
                
            self.header_count = len(headers)
            
            # Add headers
            for col_idx, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=str(header))
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                cell.fill = self.header_fill
                cell.border = self.border
            
            # NOW merge title rows to match the number of header columns
            if hasattr(self, 'title_row') and self.header_count > 1:
                end_column = get_column_letter(self.header_count)
                try:
                    self.worksheet.merge_cells(f'A{self.title_row}:{end_column}{self.title_row}')
                except:
                    pass  # Skip if merge fails
                
                if hasattr(self, 'subtitle_row') and self.subtitle_row:
                    try:
                        self.worksheet.merge_cells(f'A{self.subtitle_row}:{end_column}{self.subtitle_row}')
                    except:
                        pass  # Skip if merge fails
            
            self.current_row += 1
            
        except Exception as e:
            logger.error(f"Error adding Excel headers: {e}")
            pass

    def add_data_rows(self, data):
        """Add data rows with proper number formatting"""
        try:
            if not data:
                return
                
            for row in data:
                for col_idx, value in enumerate(row, 1):
                    cell = self.worksheet.cell(row=self.current_row, column=col_idx)
                    
                    # Handle empty values (convert N/A to empty)
                    if value in ['N/A', 'n/a', None]:
                        cell.value = None
                    # Handle negative values in brackets
                    elif isinstance(value, str) and value.startswith('(') and value.endswith(')'):
                        try:
                            # Extract numeric value and make it negative
                            numeric_value = -float(value.strip('()').replace(',', ''))
                            cell.value = numeric_value
                            cell.number_format = '_-[Red](#,##0.000_);#,##0.000'  # Accounting format for negatives
                        except ValueError:
                            cell.value = str(value)
                    # Handle regular numbers
                    elif isinstance(value, (int, float)):
                        cell.value = float(value)
                        # Format based on column type
                        if col_idx >= 5:  # Assuming numeric columns start from 5th
                            cell.number_format = '#,##0.000'
                    else:
                        # Try to convert string numbers to actual numbers
                        try:
                            # Check if it's a formatted currency string
                            clean_value = str(value).replace(',', '').replace('JOD', '').replace('%', '').strip()
                            if clean_value and clean_value.replace('.', '').replace('-', '').isdigit():
                                numeric_value = float(clean_value)
                                cell.value = numeric_value
                                if '%' in str(value):
                                    cell.number_format = '0%'
                                elif col_idx >= 5:
                                    cell.number_format = '#,##0.000'
                            else:
                                cell.value = str(value) if value is not None else None
                        except (ValueError, AttributeError):
                            cell.value = str(value) if value is not None else None
                    
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Right-align numeric columns
                    if col_idx >= 5 and isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        
                self.current_row += 1
                
        except Exception as e:
            logger.error(f"Error adding Excel data rows: {e}")
            pass

    def add_summary(self, summary_data):
        """Add summary section with proper number formatting"""
        try:
            if not summary_data:
                return
                
            # Add spacing before summary
            self.current_row += 1
            
            # Add summary title
            self.worksheet[f'A{self.current_row}'] = "Summary"
            self.worksheet[f'A{self.current_row}'].font = Font(size=14, bold=True, color="2C3E50")
            
            # Merge summary title across available columns
            if self.header_count > 1:
                end_column = get_column_letter(self.header_count)
                try:
                    self.worksheet.merge_cells(f'A{self.current_row}:{end_column}{self.current_row}')
                except:
                    pass
            
            self.current_row += 1
            
            # Add summary data
            for key, value in summary_data.items():
                if key and value is not None:
                    # Key in column A
                    key_cell = self.worksheet[f'A{self.current_row}']
                    key_cell.value = f"{key}:"
                    key_cell.font = Font(bold=True)
                    key_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    key_cell.border = self.border
                    
                    # Value in column B
                    value_cell = self.worksheet[f'B{self.current_row}']
                    
                    # Handle different value types
                    if isinstance(value, (int, float)):
                        value_cell.value = float(value)
                        value_cell.number_format = '#,##0.000'
                    elif isinstance(value, str) and value.startswith('(') and value.endswith(')'):
                        # Handle negative values in brackets
                        try:
                            numeric_value = -float(value.strip('()').replace(',', ''))
                            value_cell.value = numeric_value
                            value_cell.number_format = '_-[Red](#,##0.000_);#,##0.000'
                        except ValueError:
                            value_cell.value = str(value)
                    else:
                        # Try to parse numbers from formatted strings
                        try:
                            clean_value = str(value).replace(',', '').replace('JOD', '').replace('%', '').strip()
                            if clean_value and clean_value.replace('.', '').replace('-', '').isdigit():
                                numeric_value = float(clean_value)
                                value_cell.value = numeric_value
                                if '%' in str(value):
                                    value_cell.number_format = '0%'
                                else:
                                    value_cell.number_format = '#,##0.000'
                            else:
                                value_cell.value = str(value)
                        except (ValueError, AttributeError):
                            value_cell.value = str(value)
                    
                    value_cell.border = self.border
                    value_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    self.current_row += 1
                    
        except Exception as e:
            logger.error(f"Error adding Excel summary: {e}")
            pass

    def get_response(self, filename):
        """Generate HTTP response with Excel file"""
        try:
            # Auto-fit column widths
            for column in self.worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Clean filename
            clean_filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.'))
            if not clean_filename.endswith('.xlsx'):
                clean_filename += '.xlsx'
            
            # Create response using BytesIO buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            
            # Create HTTP response with proper content
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{clean_filename}"'
            response['Content-Length'] = len(buffer.getvalue())
                
            return response
            
        except Exception as e:
            logger.error(f"Error generating Excel response: {e}")
            # Return a JSON response instead of failing completely
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'Failed to generate Excel file: {str(e)}'})
    def add_section_header(self, header_text):
        """Add a section header with special formatting"""
        if not self.current_row:
            self.current_row = 1
        
        # Add some spacing
        self.current_row += 2
        
        # Add header with formatting
        cell = self.worksheet.cell(row=self.current_row, column=1, value=header_text)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Merge across multiple columns for better appearance
        self.worksheet.merge_cells(f"A{self.current_row}:H{self.current_row}")
        
        # Add border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
        
        self.current_row += 1

    def add_text_row(self, text_items):
        """Add a row of text items (for checklists, alerts, etc.)"""
        if not self.current_row:
            self.current_row = 1
        
        for col_idx, text in enumerate(text_items, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=text)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Special formatting for checklist items
            if text.startswith('□'):
                cell.font = Font(bold=True, size=11)
            elif text.startswith('⚠'):
                cell.font = Font(bold=True, color="DC3545")
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            elif text.startswith('✅') or text.startswith('✓'):
                cell.font = Font(bold=True, color="198754")
                cell.fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
        
        # Merge across multiple columns for full-width text
        if len(text_items) == 1:
            self.worksheet.merge_cells(f"A{self.current_row}:H{self.current_row}")
        
        self.current_row += 1

    def add_verification_table(self, headers, data, table_style="calculations"):
        """Add a specially formatted verification table"""
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=header)
            if table_style == "calculations":
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
            else:
                cell.font = Font(bold=True, color="495057")
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        self.current_row += 1
        
        # Add data rows
        for row_data in data:
            for col_idx, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Special formatting for monetary values
                if isinstance(value, str) and 'JOD' in value:
                    cell.font = Font(bold=True)
                    cell.number_format = '0.000'
            
            self.current_row += 1

    def add_verification_summary(self, total_amount):
        """Add a verification summary row"""
        # Add a prominent total verification row
        verification_row = self.current_row
        
        # Create the verification summary
        self.worksheet.cell(row=verification_row, column=1, value="TOTAL VERIFICATION")
        self.worksheet.cell(row=verification_row, column=2, value="Sum of all calculations")
        self.worksheet.cell(row=verification_row, column=3, value=f"{total_amount:.3f} JOD")
        
        # Format the entire row
        for col in range(1, 4):
            cell = self.worksheet.cell(row=verification_row, column=col)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
            cell.border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        self.current_row += 1


class PDFExporter:
    def __init__(self, title="Report", landscape_mode=False):
        """Initialize PDF exporter with proper error handling"""
        try:
            self.title = title
            self.landscape_mode = landscape_mode
            self.styles = getSampleStyleSheet()
            self.story = []
            
            # Determine page size and orientation
            if self.landscape_mode:
                self.pagesize = landscape(A4)  # Landscape A4: 842 x 595 points
                self.available_width = 842 - 2*inch  # ~650 points available
            else:
                self.pagesize = A4  # Portrait A4: 595 x 842 points
                self.available_width = 595 - 2*inch  # ~450 points available
            
            # Custom styles with better sizing
            self.title_style = ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Heading1'],
                fontSize=18,
                spaceAfter=20,
                alignment=1,  # Center alignment
                textColor=colors.HexColor('#0f4c75')
            )
            
            self.subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=self.styles['Normal'],
                fontSize=12,
                spaceAfter=15,
                alignment=1,
                textColor=colors.HexColor('#6c757d')
            )
            
            self.header_style = ParagraphStyle(
                'CustomHeader',
                parent=self.styles['Heading2'],
                fontSize=14,
                spaceAfter=10,
                textColor=colors.HexColor('#0f4c75')
            )
            
        except Exception as e:
            logger.error(f"PDFExporter initialization error: {e}")
            raise Exception(f"Failed to initialize PDF exporter: {str(e)}")
        
    def add_title(self, title, subtitle=None):
        """Add title with proper formatting"""
        try:
            self.story.append(Paragraph(str(title) if title is not None else '', self.title_style))
            if subtitle:
                self.story.append(Paragraph(str(subtitle), self.subtitle_style))
            self.story.append(Spacer(1, 20))
        except Exception as e:
            logger.error(f"Error adding PDF title: {e}")
            # Don't fail entirely, just skip title
            pass
        
    def add_metadata(self, metadata):
        """Add metadata table with better sizing"""
        try:
            if not metadata:
                return
                
            data = []
            for key, value in metadata.items():
                if key and value is not None:
                    data.append([str(key) + ":", str(value)])
            
            if not data:
                return
            
            # Calculate column widths based on available space
            col1_width = self.available_width * 0.35
            col2_width = self.available_width * 0.65
            
            table = Table(data, colWidths=[col1_width, col2_width])
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            self.story.append(table)
            self.story.append(Spacer(1, 20))
            
        except Exception as e:
            logger.error(f"Error adding PDF metadata: {e}")
            # Don't fail entirely, just skip metadata
            pass
        
    def add_table(self, headers, data, auto_size_columns=True):
        """Add table with dynamic column sizing and proper width management"""
        try:
            if not data:
                self.story.append(Paragraph("No data available", self.styles['Normal']))
                return
                
            # Prepare table data with proper string conversion
            table_data = []
            if headers:
                table_data.append([str(header) if header is not None else '' for header in headers])
            
            for row in data:
                if row:
                    # Convert all cells to strings and handle None values, truncate long text
                    string_row = []
                    for cell in row:
                        if cell is None:
                            string_row.append('')
                        else:
                            cell_str = str(cell)
                            # Truncate very long text for PDF display
                            string_row.append(cell_str[:50] + '...' if len(cell_str) > 50 else cell_str)
                    table_data.append(string_row)
            
            if not table_data:
                self.story.append(Paragraph("No data available", self.styles['Normal']))
                return
            
            if auto_size_columns and len(table_data[0]) > 0:
                # Calculate optimal column widths
                num_cols = len(table_data[0])
                col_widths = []
                
                # Calculate content-based widths
                for col_idx in range(num_cols):
                    max_content_length = 0
                    for row in table_data:
                        if col_idx < len(row):
                            content_length = len(str(row[col_idx]))
                            max_content_length = max(max_content_length, content_length)
                    
                    # Convert to actual width
                    # Base width calculation: 6 points per character + padding
                    min_width = 60  # Minimum 60 points (~0.83 inch)
                    estimated_width = max_content_length * 6 + 20  # 6 points per char + padding
                    col_widths.append(max(min_width, estimated_width))
                
                # Scale down if total width exceeds available space
                total_width = sum(col_widths)
                if total_width > self.available_width:
                    scale_factor = self.available_width / total_width
                    col_widths = [w * scale_factor for w in col_widths]
                    
                # Ensure minimum readable width
                min_readable_width = 50
                col_widths = [max(w, min_readable_width) for w in col_widths]
                
            else:
                # Equal width columns
                col_width = self.available_width / len(table_data[0])
                col_widths = [col_width] * len(table_data[0])
            
            # Create table
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Enhanced table styling
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f4c75')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Data rows styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                
                # Handle text wrapping
                ('WORDWRAP', (0, 0), (-1, -1), 'LTR'),
            ]))
            
            self.story.append(table)
            self.story.append(Spacer(1, 20))
            
        except Exception as e:
            logger.error(f"Error adding PDF table: {e}")
            self.story.append(Paragraph(f"Error loading table data: {str(e)}", self.styles['Normal']))
        
    def add_summary(self, summary_data):
        """Add summary section with better layout"""
        try:
            if not summary_data:
                return
                
            self.story.append(Paragraph("Summary", self.header_style))
            
            data = []
            for key, value in summary_data.items():
                if key and value is not None:
                    data.append([str(key) + ":", str(value)])
            
            if not data:
                return
            
            # Use smaller width for summary table (centered)
            col1_width = self.available_width * 0.4
            col2_width = self.available_width * 0.35
            
            table = Table(data, colWidths=[col1_width, col2_width])
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e3f2fd')),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            self.story.append(table)
            self.story.append(Spacer(1, 20))
            
        except Exception as e:
            logger.error(f"Error adding PDF summary: {e}")
            # Don't fail entirely, just skip summary
            pass
        
    def get_response(self, filename):
        """Generate HTTP response with properly sized PDF file - FIXED VERSION"""
        try:
            # Clean filename
            clean_filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
            if not clean_filename.endswith('.pdf'):
                clean_filename += '.pdf'
            
            # Create PDF using BytesIO buffer - THIS FIXES THE CORRUPTION
            buffer = io.BytesIO()
            
            # Create PDF with custom page size and margins
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=self.pagesize,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=1*inch,
                bottomMargin=0.75*inch
            )
            
            doc.build(self.story)
            buffer.seek(0)
            
            # Create HTTP response with proper content
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="{clean_filename}"'
            response['Content-Length'] = len(buffer.getvalue())
            
            return response
            
        except Exception as e:
            logger.error(f"Error generating PDF response: {e}")
            # Return a JSON response instead of failing completely
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'Failed to generate PDF file: {str(e)}'})


def create_pdf_exporter_for_data(title, data_width="normal"):
    """
    Helper function to create appropriately sized PDF exporter
    
    Args:
        title: PDF title
        data_width: "normal" for portrait, "wide" for landscape
    """
    try:
        landscape_mode = data_width == "wide"
        return PDFExporter(title=title, landscape_mode=landscape_mode)
    except Exception as e:
        logger.error(f"Error creating PDF exporter: {e}")
        raise Exception(f"Failed to create PDF exporter: {str(e)}")