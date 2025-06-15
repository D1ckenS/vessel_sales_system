from django.utils import timezone
from datetime import datetime, timedelta, date
from django.contrib.auth.decorators import login_required
import calendar
from django.db.models import Sum, Count, F, Q
import json
from django.template.loader import render_to_string
import weasyprint
import io
from products import models
from .utils import BilingualMessages
from django.http import JsonResponse, HttpResponse
from transactions.models import Transaction, InventoryLot, Trip, PurchaseOrder
from vessels.models import Vessel
from .utils.exports import ExcelExporter
from .utils.weasy_exporter import create_weasy_exporter_for_data, create_weasy_exporter
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404
import logging
from products.models import Product, Category
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Set up logging
logger = logging.getLogger(__name__)

# ===============================================================================
# HELPER FUNCTIONS
# ===============================================================================

def format_currency(value, decimals=3):
    """Format currency with specified decimal places"""
    try:
        return f"{float(value):.{decimals}f}"
    except (ValueError, TypeError):
        return "0.000"

def format_currency_or_none(value, decimals=3):
    """Format currency or return None for empty/zero values"""
    try:
        float_val = float(value)
        if float_val == 0:
            return None  # Return None instead of "0.000"
        return f"{float_val:.{decimals}f}"
    except (ValueError, TypeError):
        return None
    
def format_percentage(value):
    """Format percentage with 0 decimal places"""
    try:
        return f"{float(value):.0f}%"
    except (ValueError, TypeError):
        return "0%"

def format_negative_if_supply(value, transaction_type):
    """Format value as negative if it's a supply transaction"""
    try:
        formatted_value = float(value)
        if transaction_type in ['SUPPLY', 'TRANSFER_IN']:
            if formatted_value != 0:  # Only bracket non-zero values
                return f"({format_currency(abs(formatted_value), 3)})"
        return format_currency(abs(formatted_value), 3)
    except (ValueError, TypeError):
        return None  # Return None instead of "0.000" for empty values

def get_fifo_cost_for_transfer(vessel, product, quantity):
    """Get FIFO cost for transfer without actually consuming inventory"""
    lots = InventoryLot.objects.filter(
        vessel=vessel,
        product=product,
        remaining_quantity__gt=0
    ).order_by('purchase_date', 'created_at')
    
    remaining_to_transfer = float(quantity)
    total_cost = 0
    
    for lot in lots:
        if remaining_to_transfer <= 0:
            break
            
        available = float(lot.remaining_quantity)
        to_take = min(available, remaining_to_transfer)
        
        total_cost += to_take * float(lot.purchase_price)
        remaining_to_transfer -= to_take
    
    if remaining_to_transfer > 0:
        # Not enough inventory, use product's purchase price as fallback
        if hasattr(product, 'purchase_price'):
            total_cost += remaining_to_transfer * float(product.purchase_price)
        else:
            total_cost += remaining_to_transfer * float(product.cost_price or 0)
    
    return total_cost / float(quantity) if quantity > 0 else 0

def calculate_transfer_amounts(transaction):
    """Calculate proper amounts for transfer transactions using FIFO"""
    if transaction.transaction_type == 'TRANSFER_OUT':
        # Use FIFO cost as "revenue" for transfer out
        try:
            fifo_cost = get_fifo_cost_for_transfer(
                transaction.vessel, 
                transaction.product, 
                transaction.quantity
            )
            return fifo_cost * float(transaction.quantity)
        except Exception as e:
            logger.warning(f"Error calculating FIFO cost for transfer: {e}")
            return float(transaction.quantity) * float(transaction.unit_price)
    elif transaction.transaction_type == 'TRANSFER_IN':
        # Use the same FIFO cost as "supply cost" for transfer in
        try:
            if hasattr(transaction, 'related_transfer') and transaction.related_transfer:
                return calculate_transfer_amounts(transaction.related_transfer)
            else:
                # Fallback: use the transaction's unit price
                return float(transaction.quantity) * float(transaction.unit_price)
        except Exception as e:
            logger.warning(f"Error calculating transfer in cost: {e}")
            return float(transaction.quantity) * float(transaction.unit_price)
    return float(transaction.quantity) * float(transaction.unit_price)

def calculate_totals_by_type(transactions):
    """Calculate totals split by transaction type"""
    totals = {
        'total_sales': 0,
        'total_supplies': 0,
        'total_transfers': 0
    }
    
    for transaction in transactions:
        if transaction.transaction_type == 'SALE':
            totals['total_sales'] += float(transaction.quantity) * float(transaction.unit_price)
        elif transaction.transaction_type == 'SUPPLY':
            totals['total_supplies'] += float(transaction.quantity) * float(transaction.unit_price)
        elif transaction.transaction_type in ['TRANSFER_OUT', 'TRANSFER_IN']:
            # Only count transfer out to avoid double counting
            if transaction.transaction_type == 'TRANSFER_OUT':
                totals['total_transfers'] += calculate_transfer_amounts(transaction)
    
    return totals

def calculate_product_level_summary(transactions):
    """Calculate product-level summary for reports"""
    
    
    products = defaultdict(lambda: {
        'name': '',
        'product_id': '',
        'qty_supplied': 0,
        'qty_sold': 0, 
        'total_cost': 0,
        'total_revenue': 0
    })
    
    for transaction in transactions:
        if not transaction.product:
            continue
            
        product_key = transaction.product.id
        product_data = products[product_key]
        
        # Set product info
        product_data['name'] = transaction.product.name
        product_data['product_id'] = getattr(transaction.product, 'product_id', 'N/A')
        
        # Calculate quantities and amounts
        quantity = safe_float(transaction.quantity)
        
        if transaction.transaction_type == 'SUPPLY':
            product_data['qty_supplied'] += quantity
            product_data['total_cost'] += quantity * safe_float(transaction.unit_price)
        elif transaction.transaction_type == 'SALE':
            product_data['qty_sold'] += quantity
            product_data['total_revenue'] += quantity * safe_float(transaction.unit_price)
        # Note: Transfers are not included in product summary as they don't change overall inventory
    
    # Convert to list format for export
    summary_data = []
    for product_data in products.values():
        net_profit = product_data['total_revenue'] - product_data['total_cost']
        summary_data.append([
            product_data['name'],
            product_data['product_id'],
            format_currency(product_data['qty_supplied'], 3),
            format_currency(product_data['qty_sold'], 3),
            format_currency(product_data['total_cost'], 3),
            format_currency(product_data['total_revenue'], 3),
            format_currency(net_profit, 3)
        ])
    
    # Sort by net profit descending
    summary_data.sort(key=lambda x: float(x[6].replace(',', '').replace('(', '-').replace(')', '')), reverse=True)
    
    return summary_data

def get_translated_labels(request, data=None):
    """Get translated labels based on user's language preference"""
    
    # Try to get language from request data first (sent by JavaScript)
    user_language = 'en'
    if data and 'language' in data:
        user_language = data.get('language', 'en')
    else:
        # Fallback to session
        user_language = request.session.get('preferred_language', 'en')
    
    # Ensure it's a valid language
    if user_language not in ['en', 'ar']:
        user_language = 'en'
    
    # Define translations
    translations = {
        'en': {
            # Basic Labels
            'export_date': 'Export Date',
            'trip_number': 'Trip Number',
            'vessel': 'Vessel',
            'trip_date': 'Trip Date',
            'status': 'Status',
            'passengers': 'Passengers',
            'total_revenue_jod': 'Total Revenue',
            'total_cogs_jod': 'Sales Purchase Price',
            'total_profit_jod': 'Sales - Sales Purchase Price',
            'profit_margin': 'Profit Margin',
            'generated_by': 'Generated By',
            'time': 'Time',
            'product': 'Product',
            'product_id': 'Product ID',
            'quantity': 'Quantity',
            'unit_selling_price_jod': 'Unit Selling Price',
            'cogs_per_unit_jod': 'Purchase price/Unit',
            'revenue_jod': 'Revenue',
            'cogs_jod': 'Sales Purchase Price',
            'profit_jod': 'Sales - Sales Purchase Price',
            'notes': 'Notes',
            'total_items_sold': 'Total Items Sold',
            'completed': 'Completed',
            'pending': 'Pending',
            'supplies': 'Supplies',
            
            'period': 'Period',
            'total_trips': 'Total Trips',
            'trips_overview': 'Trips Overview',
            'total_pos': 'Total Number of Purchase Orders',
            'pos_overview': 'Purchase Orders Overview',
            'total_quantity': 'Total Quantity',
            'currency': 'Currency',
            'jod': 'JOD',
            
            # Report Titles
            'trip_sales_report': 'Trip Sales Report',
            'trips_report': 'Trips Sales Report',
            'purchase_order_supply_report': 'Purchase Order Supply Report',
            
            # Template Elements
            'generated_on': 'Generated on',
            'report_information': 'Report Information',
            'summary': 'Summary',
            'company_logo': 'COMPANY LOGO',
            'no_data_available': 'No data available',
            'transaction_details': 'Transaction Details',
            'item_details': 'Item Details',
            
            # PO Specific
            'po_number': 'PO Number',
            'po_date': 'PO Date',
            'unit_cost': 'Purchase price/unit',
            'total_cost_jod': 'Total Cost (JOD)',
            'total_items_received': 'Total Items Received',
            'average_cost_per_item_jod': 'Average Cost per Item (JOD)',
            
            # Report Titles with Dynamic Content
            'trip_report_title': 'Trip {trip_number} Report',
            'po_report_title': 'PO {po_number} Report',
            
            # Inventory Export
            'inventory_report': 'Inventory Report',
            'current_inventory': 'Current Inventory Status',
            'product_name': 'Product Name',
            'category': 'Category',
            'current_stock': 'Current Stock',
            'minimum_level': 'Minimum Level',
            'unit_cost': 'Unit Cost',
            'total_value_jod': 'Total Value',
            'last_updated': 'Last Updated',
            'total_products': 'Total Products',
            'low_stock_items': 'Low Stock Items',
            'average_value_per_item': 'Average Value per Item',
            'uncategorized': 'Uncategorized',
            'all': 'All',
            'system': 'System',
            'yes': 'Yes',
            'no': 'No',
            'low_stock_filter': 'Low Stock Only',
            'total_value': 'Total Value',
            
            # Transactions Export  
            'transactions_report': 'Transactions Report',
            'transaction_history': 'Transaction History',
            'date': 'Date',
            'type': 'Type',
            'unit_price_jod': 'Unit Price',
            'total_amount_jod': 'Total Amount',
            'created_by': 'Created By',
            'total_records': 'Total Records',
            'total_sales_jod': 'Total Sales',
            'total_supplies_jod': 'Total Supplies',
            'total_transfers_jod': 'Total Transfers',
            
            # Report Exports
            'monthly_report': 'Monthly Report',
            'daily_report': 'Daily Report',
            'analytics_report': 'Analytics Report',
            'analysis_period': 'Analysis Period',
            'vessel_performance': 'Vessel Performance',
            'revenue_jod': 'Revenue',
            'costs_jod': 'Costs',
            'profit_margin_percent': 'Profit Margin (%)',
            'sales_count': 'Sales Count',
            'avg_transaction_value': 'Avg Transaction Value',
            'po_count': 'PO Count',
            'best_performing_vessel': 'Best Performing Vessel',
            'daily_performance': 'Daily Performance',
            'average_per_trip': 'Average per Trip',
            'transactions': 'Transactions',
        },
        'ar': {
            # Basic Labels (Arabic)
            'export_date': 'تاريخ التصدير',
            'trip_number': 'رقم الرحلة',
            'vessel': 'السفينة',
            'trip_date': 'تاريخ الرحلة',
            'status': 'الحالة',
            'passengers': 'الركاب',
            'total_revenue_jod': 'إجمالي الإيرادات',
            'total_cogs_jod': 'سعر شراء المبيعات',
            'total_profit_jod': 'المبيعات - سعر شراء المبيعات',
            'profit_margin': 'هامش الربح',
            'generated_by': 'تم إنشاؤه بواسطة',
            'time': 'الوقت',
            'product': 'المنتج',
            'product_id': 'رقم المنتج',
            'quantity': 'الكمية',
            'unit_selling_price_jod': 'سعر البيع للوحدة',
            'cogs_per_unit_jod': 'سعر شراء للوحدة',
            'revenue_jod': 'الإيرادات',
            'cogs_jod': 'سعر شراء المبيعات',
            'profit_jod': 'المبيعات - سعر شراء المبيعات',
            'notes': 'ملاحظات',
            'total_items_sold': 'إجمالي العناصر المباعة',
            'completed': 'مكتمل',
            'pending': 'معلق',
            'supplies': 'توريدات',
            
            'period': 'الفترة',
            'total_trips': 'إجمالي الرحلات',
            'trips_overview': 'نظرة عامة على الرحلات',
            'total_pos': 'عدد أوامر الشراء',
            'pos_overview': 'نظرة عامة على أوامر الشراء',
            'total_quantity': 'إجمالي الكمية',
            'currency': 'العملة',
            'jod': 'دينار',
            
            # Report Titles (Arabic)
            'trip_sales_report': 'تقرير مبيعات الرحلة',
            'trips_report': 'تقرير مبيعات الرحلات',
            'purchase_order_supply_report': 'تقرير أوامر الشراء',
            
            # Template Elements (Arabic)
            'generated_on': 'تم إنشاؤه في',
            'report_information': 'معلومات التقرير',
            'summary': 'الملخص',
            'company_logo': 'شعار الشركة',
            'no_data_available': 'لا توجد بيانات متاحة',
            'transaction_details': 'تفاصيل المعاملات',
            'item_details': 'تفاصيل العناصر',
            
            # PO Specific (Arabic)
            'po_number': 'رقم أمر الشراء',
            'po_date': 'تاريخ أمر الشراء',
            'unit_cost': 'تكلفة الوحدة (دينار)',
            'total_cost_jod': 'التكلفة الإجمالية (دينار)',
            'total_items_received': 'إجمالي العناصر المستلمة',
            'average_cost_per_item_jod': 'متوسط التكلفة لكل عنصر (دينار)',
            
            # Report Titles with Dynamic Content (Arabic)
            'trip_report_title': 'تقرير رحلة {trip_number}',
            'po_report_title': 'تقرير أمر شراء {po_number}',
            
            # Inventory Export
            'inventory_report': 'تقرير المخزون',
            'current_inventory': 'حالة المخزون الحالية',
            'product_name': 'اسم المنتج',
            'category': 'الفئة',
            'current_stock': 'المخزون الحالي',
            'minimum_level': 'الحد الأدنى',
            'unit_cost': 'تكلفة الوحدة',
            'total_value_jod': 'القيمة الإجمالية',
            'last_updated': 'آخر تحديث',
            'total_products': 'إجمالي المنتجات',
            'low_stock_items': 'العناصر منخفضة المخزون',
            'average_value_per_item': 'متوسط القيمة لكل عنصر',
            'uncategorized': 'Uncategorized',
            'uncategorized': 'غير مصنف',
            'all': 'الكل',
            'system': 'النظام',
            'yes': 'نعم',
            'no': 'لا',
            'low_stock_filter': 'المخزون المنخفض فقط',
            'total_value': 'القيمة الإجمالية',
            
            # Transactions Export
            'transactions_report': 'تقرير المعاملات',
            'transaction_history': 'تاريخ المعاملات',
            'date': 'التاريخ',
            'type': 'النوع',
            'unit_price_jod': 'سعر الوحدة',
            'total_amount_jod': 'المبلغ الإجمالي',
            'created_by': 'تم إنشاؤه بواسطة',
            'total_records': 'إجمالي السجلات',
            'total_sales_jod': 'إجمالي المبيعات',
            'total_supplies_jod': 'إجمالي التوريدات',
            'total_transfers_jod': 'إجمالي التحويلات',
            
            # Report Exports
            'monthly_report': 'تقرير شهري',
            'daily_report': 'تقرير يومي',
            'analytics_report': 'تقرير تحليلي',
            'analysis_period': 'فترة التحليل',
            'vessel_performance': 'أداء السفن',
            'revenue_jod': 'الإيرادات',
            'costs_jod': 'التكاليف',
            'profit_margin_percent': 'هامش الربح (%)',
            'sales_count': 'عدد المبيعات',
            'avg_transaction_value': 'متوسط قيمة المعاملة',
            'po_count': 'عدد أوامر الشراء',
            'best_performing_vessel': 'أفضل سفينة أداءً',
            'daily_performance': 'الأداء اليومي',
            'average_per_trip': 'متوسط لكل رحلة',
            'transactions': 'المعاملات',
        }
    }
    
    labels = translations.get(user_language, translations['en'])
    labels['language'] = user_language  # Add language info
    return labels

def translate_numbers_to_arabic(text, language):
    """Convert Western numerals to Arabic-Indic numerals if language is Arabic"""
    if language != 'ar':
        return text
    
    # Arabic-Indic numerals mapping
    arabic_numerals = {
        '0': '٠', '1': '١', '2': '٢', '3': '٣', '4': '٤',
        '5': '٥', '6': '٦', '7': '٧', '8': '٨', '9': '٩'
    }
    
    # Convert numbers in text
    result = str(text)
    for western, arabic in arabic_numerals.items():
        result = result.replace(western, arabic)
    
    return result

def get_vessel_name_by_language(vessel, language):
    """Get vessel name in appropriate language"""
    if language == 'ar' and vessel and hasattr(vessel, 'name_ar') and vessel.name_ar:
        return vessel.name_ar
    return vessel.name if vessel else 'N/A'

def format_date_for_language(date_obj, language):
    """Format date according to language preference"""
    if not date_obj:
        return ''
    
    if language == 'ar':
        # Arabic date format: DD/MM/YYYY with Arabic numerals
        formatted = date_obj.strftime('%d/%m/%Y')
        return translate_numbers_to_arabic(formatted, 'ar')
    else:
        # English date format
        return date_obj.strftime('%d/%m/%Y')

def format_datetime_for_language(datetime_obj, language):
    """Format datetime according to language preference"""
    if not datetime_obj:
        return ''
    
    if language == 'ar':
        # Arabic datetime format with Arabic numerals
        formatted = datetime_obj.strftime('%d/%m/%Y %H:%M')
        return translate_numbers_to_arabic(formatted, 'ar')
    else:
        # English datetime format
        return datetime_obj.strftime('%d/%m/%Y %H:%M')

def safe_float(value, default=0.0):
    """Safely convert value to float"""
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default

def safe_int(value, default=0):
    """Safely convert value to int"""
    try:
        return int(value) if value is not None else default
    except (ValueError, TypeError):
        return default

def format_date(dt):
    """Format date for export"""
    if dt:
        return dt.strftime('%d/%m/%Y')
    return ''

def get_date_range_from_request(data):
    """Extract and validate date range from request data"""
    try:
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date() - timedelta(days=30)
            
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = timezone.now().date()
            
        return start_date, end_date
    except ValueError as e:
        logger.error(f"Date parsing error: {e}")
        return None, None

def create_safe_response(content, content_type, filename):
    """Create a safe HTTP response with proper headers"""
    response = HttpResponse(content_type=content_type)
    clean_filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
    response['Content-Disposition'] = f'attachment; filename="{clean_filename}"'
    response['Content-Length'] = len(content)
    return response

# ===============================================================================
# INVENTORY EXPORT
# ===============================================================================

@login_required
@require_http_methods(["POST"])
def export_inventory(request):
    """Export current inventory status to Excel or PDF - Enhanced with RTL support"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get translated labels with language from request
        labels = get_translated_labels(request, data)
        language = labels['language']
        
        # Get filters
        vessel_id = data.get('vessel_id')  # Add vessel filter support
        category_id = data.get('category_id')
        low_stock_only = data.get('low_stock_only', False)
        
        # Build query for inventory lots
        inventory_lots = InventoryLot.objects.select_related(
            'product', 'product__category', 'vessel'
        ).filter(
            remaining_quantity__gt=0  # Only active inventory with remaining stock
        ).order_by('product__name', 'vessel__name')
        
        # Apply filters
        if vessel_id:
            inventory_lots = inventory_lots.filter(vessel_id=vessel_id)
        if category_id:
            inventory_lots = inventory_lots.filter(product__category_id=category_id)
        if low_stock_only:
            inventory_lots = inventory_lots.filter(
                remaining_quantity__lte=F('product__minimum_stock_level')
            )
        
        # Prepare inventory data
        inventory_data = []
        total_value = 0
        low_stock_count = 0
        
        for lot in inventory_lots[:2000]:  # Limit to prevent memory issues
            # Get unit cost (use purchase_price from lot or product cost_price as fallback)
            unit_cost = safe_float(lot.purchase_price if hasattr(lot, 'purchase_price') else 
                                 (lot.product.cost_price if lot.product else 0))
            
            # Calculate total cost for this lot
            total_cost = safe_float(lot.remaining_quantity) * unit_cost
            total_value += total_cost
            
            # Check if low stock
            min_level = safe_float(getattr(lot.product, 'minimum_stock_level', 0)) if lot.product else 0
            if lot.remaining_quantity <= min_level:
                low_stock_count += 1
            
            # Get vessel name in appropriate language
            vessel_name = get_vessel_name_by_language(lot.vessel, language)
            
            # Format data with RTL number translation
            formatted_data = [
                lot.product.name if lot.product else 'N/A',
                translate_numbers_to_arabic(lot.product.item_id, language) if lot.product else 'N/A',
                lot.product.category.name if lot.product and lot.product.category else 'N/A',
                vessel_name,
                translate_numbers_to_arabic(format_currency(lot.remaining_quantity, 3), language),
                translate_numbers_to_arabic(format_currency(min_level, 3), language),
                translate_numbers_to_arabic(format_currency(unit_cost, 3), language),
                translate_numbers_to_arabic(format_currency(total_cost, 3), language),
                format_date_for_language(lot.purchase_date if hasattr(lot, 'purchase_date') else 
                                       (lot.created_at.date() if hasattr(lot, 'created_at') else datetime.now().date()), language)
            ]
            inventory_data.append(formatted_data)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"inventory_export_{timestamp}"
        
        # Get current datetime for metadata
        from django.utils import timezone
        local_dt = timezone.localtime(timezone.now()).replace(tzinfo=None)
        
        # Metadata with translation and currency pattern
        metadata = {
            labels['export_date']: format_datetime_for_language(local_dt, language),
            labels['total_products']: translate_numbers_to_arabic(str(len(inventory_data)), language),
            labels['total_value']: translate_numbers_to_arabic(format_currency(total_value, 3), language),
            labels['currency']: labels['jod'],  # Apply new currency pattern
            labels['low_stock_items']: translate_numbers_to_arabic(str(low_stock_count), language),
            labels['average_value_per_item']: translate_numbers_to_arabic(format_currency((total_value / len(inventory_data)) if inventory_data else 0, 3), language),
            labels['generated_by']: request.user.username,
            labels['vessel']: vessel_id or labels.get('all', 'All'),
            labels.get('category', 'Category'): category_id or labels.get('all', 'All'),
            labels.get('filter_type', 'Filter'): labels['low_stock_only'] if low_stock_only else labels.get('all_items', 'All Items')
        }
        
        # Headers with translation (no JOD/دينار in headers)
        headers = [
            labels['product_name'],
            labels['product_id'],
            labels['category'],
            labels['vessel'],
            labels['current_stock'],
            labels['minimum_level'],
            labels['unit_cost'],  # No (JOD) suffix
            labels['total_value'],  # No (JOD) suffix
            labels['last_updated']
        ]
        
        # Summary data with translation and currency pattern
        summary_data = {
            labels['total_products']: translate_numbers_to_arabic(str(len(inventory_data)), language),
            labels['total_value']: translate_numbers_to_arabic(format_currency(total_value, 3), language),
            labels['low_stock_items']: translate_numbers_to_arabic(str(low_stock_count), language),
            labels['average_value_per_item']: translate_numbers_to_arabic(format_currency((total_value / len(inventory_data)) if inventory_data else 0, 3), language),
            labels['currency']: labels['jod']  # Apply new currency pattern
        }
        
        # Generate report title with translation
        report_title = labels['inventory_report']
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title=report_title)
                
                generation_text = f"{labels['generated_on']} {format_datetime_for_language(local_dt, language)}"
                exporter.add_title(report_title, generation_text)
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(inventory_data)
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel inventory export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF - Use same pattern as working individual exports
            try:
                # Create context manually like the working individual exports
                context = {
                    'title': report_title,
                    'metadata': metadata,
                    'tables': [{'title': f"{report_title} - {labels['current_inventory']}", 'id': 'inventory_table', 'headers': headers, 'rows': inventory_data}],
                    'charts': [],
                    'summary_data': summary_data,
                    'orientation': 'landscape',
                    'language': language,  # This enables all RTL styling
                    'generation_date': format_datetime_for_language(local_dt, language),
                    'has_logo': False,
                    'generated_on_text': labels['generated_on'],
                    'report_info_text': labels['report_information'],
                    'summary_text': labels['summary'],
                    'company_logo_text': labels['company_logo'],
                    'no_data_text': labels['no_data_available'],
                }
                
                # Use same pattern as working individual exports
                from django.template.loader import render_to_string
                import weasyprint
                import io
                
                template_name = 'frontend/exports/wide_report.html'
                html_string = render_to_string(template_name, context)
                
                html = weasyprint.HTML(string=html_string)
                
                buffer = io.BytesIO()
                html.write_pdf(target=buffer)
                buffer.seek(0)
                
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                response['Content-Length'] = len(buffer.getvalue())
                response.write(buffer.getvalue())
                
                return response
                
            except Exception as e:
                logger.error(f"PDF inventory export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Inventory export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

# ===============================================================================
# TRANSACTION EXPORTS
# ===============================================================================

@login_required
@require_http_methods(["POST"])
def export_transactions(request):
    """Export transactions to Excel or PDF with comprehensive filtering"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get filters
        start_date, end_date = get_date_range_from_request(data)
        if not start_date or not end_date:
            return JsonResponse({'success': False, 'error': 'Invalid date range'})
            
        vessel_id = data.get('vessel_id')
        transaction_type = data.get('transaction_type')
        product_id = data.get('product_id')
        
        # Build query with proper field names
        transactions = Transaction.objects.select_related(
            'vessel', 'product', 'product__category', 'created_by', 'trip', 'purchase_order'
        ).filter(
            transaction_date__gte=start_date,
            transaction_date__lte=end_date
        ).order_by('-transaction_date')
        
        # Apply filters
        if vessel_id:
            transactions = transactions.filter(vessel_id=vessel_id)
        if transaction_type:
            transactions = transactions.filter(transaction_type=transaction_type)
        if product_id:
            transactions = transactions.filter(product_id=product_id)
            
        # Calculate totals by type
        transaction_list = list(transactions[:5000])  # Limit to prevent memory issues
        totals_by_type = calculate_totals_by_type(transaction_list)
        
        # Prepare table data with new formatting
        table_data = []
        
        for transaction in transaction_list:
            # Calculate amounts based on transaction type
            if transaction.transaction_type in ['TRANSFER_OUT', 'TRANSFER_IN']:
                amount = calculate_transfer_amounts(transaction)
            else:
                amount = safe_float(transaction.quantity) * safe_float(transaction.unit_price)
            
            # Determine Unit Cost and Unit Price based on transaction type
            unit_cost = None
            unit_price = None
            
            if transaction.transaction_type == 'SALE':
                unit_price = safe_float(transaction.unit_price)
                # For sales, unit cost would be the FIFO cost (if available)
                # For now, we'll leave unit_cost as None for sales
            elif transaction.transaction_type == 'SUPPLY':
                unit_cost = safe_float(transaction.unit_price)
                # For supplies, unit_price is None
            elif transaction.transaction_type == 'TRANSFER_OUT':
                # For transfer out, use FIFO cost as unit_cost
                fifo_cost = get_fifo_cost_for_transfer(transaction.vessel, transaction.product, transaction.quantity)
                unit_cost = fifo_cost
            elif transaction.transaction_type == 'TRANSFER_IN':
                # For transfer in, use the same cost as the related transfer out
                if hasattr(transaction, 'related_transfer') and transaction.related_transfer:
                    fifo_cost = get_fifo_cost_for_transfer(transaction.related_transfer.vessel, transaction.product, transaction.quantity)
                    unit_cost = fifo_cost
            
            # Format amounts (negative for supplies and transfer ins)
            formatted_amount = format_negative_if_supply(amount, transaction.transaction_type)
            
            # Format vessel, product names (remove N/A)
            vessel_name = transaction.vessel.name if transaction.vessel else None
            product_name = transaction.product.name if transaction.product else None
            category_name = transaction.product.category.name if transaction.product and transaction.product.category else None
            trip_number = transaction.trip.trip_number if transaction.trip else None
            po_number = transaction.purchase_order.po_number if transaction.purchase_order else None
            created_by = transaction.created_by.username if transaction.created_by else None
            notes = transaction.notes if transaction.notes else None
            
            table_data.append([
                format_date(transaction.transaction_date),
                transaction.get_transaction_type_display() if hasattr(transaction, 'get_transaction_type_display') else transaction.transaction_type,
                vessel_name,
                product_name,
                category_name,
                format_currency(transaction.quantity, 3),
                format_currency_or_none(unit_price, 3),  # Unit Price
                format_currency_or_none(unit_cost, 3),   # Unit Cost  
                formatted_amount,  # Total Amount (negative for supplies)
                trip_number,
                po_number,
                created_by,
                notes
            ])
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"transactions_export_{timestamp}"
        
        # Metadata with split totals
        metadata = {
            'Export Date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'Date Range': f"{format_date(start_date)} to {format_date(end_date)}",
            'Total Records': len(table_data),
            'Total Sales (JOD)': format_currency(totals_by_type['total_sales'], 3),
            'Total Supplies (JOD)': f"({format_currency(totals_by_type['total_supplies'], 3)})",
            'Total Transfers (JOD)': format_currency(totals_by_type['total_transfers'], 3),
            'Generated By': request.user.username,
            'Filters Applied': f"Vessel: {vessel_id or 'All'}, Type: {transaction_type or 'All'}, Product: {product_id or 'All'}"
        }
        
        headers = [
            'Date', 'Type', 'Vessel', 'Product', 'Category', 
            'Quantity', 'Unit Price (JOD)', 'Unit Cost (JOD)', 'Total Amount (JOD)', 
            'Trip #', 'PO #', 'Created By', 'Notes'
        ]
        
        # Create product-level summary (replaces redundant summary)
        product_summary = calculate_product_level_summary(transaction_list)
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title="Transactions Export")
                exporter.add_title("Transactions Report", f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(table_data)
                
                # Only add product summary if we have data and it's not the same as metadata
                if product_summary and len(product_summary) > 0:
                    # Add spacing
                    exporter.current_row += 1
                    
                    # Add product summary title
                    exporter.worksheet[f'A{exporter.current_row}'] = "Product Summary"
                    exporter.worksheet[f'A{exporter.current_row}'].font = Font(size=14, bold=True, color="2C3E50")
                    
                    # Merge across all columns
                    if exporter.header_count > 1:
                        end_column = get_column_letter(exporter.header_count)
                        try:
                            exporter.worksheet.merge_cells(f'A{exporter.current_row}:{end_column}{exporter.current_row}')
                        except:
                            pass
                    
                    exporter.current_row += 1
                    
                    # Add product summary headers
                    summary_headers = ['Item Name', 'Item ID', 'Qty Supplied', 'Qty Sold', 'Total Cost', 'Total Revenue', 'Net Profit']
                    for col_idx, header in enumerate(summary_headers, 1):
                        cell = exporter.worksheet.cell(row=exporter.current_row, column=col_idx, value=header)
                        cell.font = exporter.header_font
                        cell.alignment = exporter.header_alignment
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                        cell.border = exporter.border
                    
                    exporter.current_row += 1
                    
                    # Add product summary data
                    for row in product_summary:
                        for col_idx, value in enumerate(row, 1):
                            cell = exporter.worksheet.cell(row=exporter.current_row, column=col_idx, value=value)
                            cell.border = exporter.border
                            if col_idx >= 3:  # Numeric columns
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                        exporter.current_row += 1
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF
            try:
                exporter = create_weasy_exporter_for_data("Transactions Report", "wide")
                exporter.add_metadata(metadata)
                exporter.add_table(headers, table_data, table_title="Transaction History")
                
                # Add product summary instead of redundant summary
                if product_summary and len(product_summary) > 0:
                    summary_headers = ['Item Name', 'Item ID', 'Qty Supplied', 'Qty Sold', 'Total Cost', 'Total Revenue', 'Net Profit']
                    exporter.add_table(summary_headers, product_summary, table_title="Product Summary")
                
                return exporter.get_response(f"{filename_base}.pdf")
                
            except Exception as e:
                logger.error(f"PDF export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Transaction export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

# ===============================================================================
# TRIP EXPORTS (LIST AND INDIVIDUAL)
# ===============================================================================

@login_required
@require_http_methods(["POST"])
def export_trips(request):
    """Export trips list to Excel or PDF - Fixed with RTL support following individual export pattern"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get translated labels with language from request (for RTL support)
        labels = get_translated_labels(request, data)
        language = labels['language']
        
        # Get filters
        start_date, end_date = get_date_range_from_request(data)
        if not start_date or not end_date:
            return JsonResponse({'success': False, 'error': 'Invalid date range'})
            
        vessel_id = data.get('vessel_id')
        status = data.get('status')
        
        # Build query
        trips = Trip.objects.select_related(
            'vessel', 'created_by'
        ).filter(
            trip_date__gte=start_date,
            trip_date__lte=end_date
        ).order_by('-trip_date')
        
        # Apply filters
        if vessel_id:
            trips = trips.filter(vessel_id=vessel_id)
        if status:
            trips = trips.filter(status=status)
            
        # Prepare data
        table_data = []
        total_revenue = 0
        
        for trip in trips[:2000]:
            # Calculate trip financials
            trip_revenue = Transaction.objects.filter(
                trip=trip,
                transaction_type='SALE'
            ).aggregate(
                total=Sum(F('quantity') * F('unit_price'))
            )['total'] or 0
            
            total_revenue += trip_revenue
            
             # Get transaction count
            transaction_count = Transaction.objects.filter(trip=trip).count()
                        
            # Get vessel name in appropriate language
            vessel_name = get_vessel_name_by_language(trip.vessel, language)
            
            # Format data with number translation for Arabic
            trip_status_key = 'completed' if getattr(trip, 'is_completed', False) or getattr(trip, 'status', '') == 'completed' else 'pending'
            trip_status = labels[trip_status_key]
            
            # Get current datetime for trip date metadata formatting
            local_dt = timezone.localtime(trip.created_at).replace(tzinfo=None) if hasattr(trip, 'created_at') else timezone.now().replace(tzinfo=None)
            
            formatted_data = [
                translate_numbers_to_arabic(str(trip.trip_number), language),
                vessel_name,
                format_date_for_language(local_dt, language),
                trip_status,
                translate_numbers_to_arabic(format_currency(trip_revenue, 3), language),
                translate_numbers_to_arabic(str(safe_int(transaction_count)), language),
                translate_numbers_to_arabic(str(safe_int(getattr(trip, 'passenger_count', 0))), language),
                trip.created_by.username if trip.created_by else 'N/A'
            ]
            table_data.append(formatted_data)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"trips_export_{timestamp}"
        
        # Metadata with translation
        metadata = {
            labels['export_date']: format_datetime_for_language(datetime.now(), language),
            labels['period']: f"{format_date_for_language(start_date, language)} - {format_date_for_language(end_date, language)}",
            labels['total_trips']: translate_numbers_to_arabic(str(len(table_data)), language),
            labels['total_revenue_jod']: translate_numbers_to_arabic(format_currency(total_revenue, 3), language),
            labels['currency']: labels['jod'],  # Apply new currency pattern
            labels['average_per_trip']: translate_numbers_to_arabic(format_currency((total_revenue / len(table_data)) if table_data else 0, 3), language),
            labels['generated_by']: request.user.username
        }
        
        # Headers with translation
        headers = [
            labels['trip_number'],
            labels['vessel'],
            labels['trip_date'],
            labels['status'],
            labels['total_revenue_jod'],
            labels.get('transactions', 'Transactions'),
            labels['passengers'],
            labels['generated_by']
        ]
        
        # Summary data with translation
        summary_data = {
            labels['total_trips']: translate_numbers_to_arabic(str(len(table_data)), language),
            labels['total_revenue_jod']: translate_numbers_to_arabic(format_currency(total_revenue, 3), language),
            labels['average_per_trip']: translate_numbers_to_arabic(format_currency((total_revenue / len(table_data)) if table_data else 0, 3), language),
            labels['currency']: labels['jod']
        }
        
        # Generate report title with translation
        if language == 'ar':
            report_title = labels['trips_report']
        else:
            report_title = labels['trips_report']
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title=report_title)
                
                generation_text = f"{labels['generated_on']} {format_datetime_for_language(local_dt, language)}"
                exporter.add_title(report_title, generation_text)
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(table_data)
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel trips export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF - Use the SAME pattern as working individual exports
            try:
                # Create context manually like the working individual exports
                context = {
                    'title': report_title,
                    'metadata': metadata,
                    'tables': [{'title': f"{report_title} - {labels.get('trips_overview', 'Trips Overview')}", 'id': 'trips_table', 'headers': headers, 'rows': table_data}],
                    'summary_data': summary_data,
                    'orientation': 'landscape',
                    'language': language,  # KEY: This enables RTL
                    'generation_date': format_datetime_for_language(datetime.now(), language),
                    'has_logo': False,
                    'generated_on_text': labels['generated_on'],
                    'report_info_text': labels['report_information'],
                    'summary_text': labels['summary'],
                    'company_logo_text': labels['company_logo'],
                    'no_data_text': labels['no_data_available'],
                }
                
                # Use the same pattern as working individual exports
                
                
                template_name = 'frontend/exports/wide_report.html'
                html_string = render_to_string(template_name, context)
                
                html = weasyprint.HTML(string=html_string)
                
                buffer = io.BytesIO()
                html.write_pdf(target=buffer)
                buffer.seek(0)
                
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                response['Content-Length'] = len(buffer.getvalue())
                response.write(buffer.getvalue())
                
                return response
                
            except Exception as e:
                logger.error(f"PDF trips export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Trips export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def export_single_trip(request, trip_id):
    """Export individual trip details - Enhanced RTL support for metadata and summary sections"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get translated labels with language from request
        labels = get_translated_labels(request, data)
        language = labels['language']
        
        # Get trip
        trip = get_object_or_404(Trip, id=trip_id)
        
        # Get trip transactions
        transactions = Transaction.objects.filter(
            trip=trip
        ).select_related('product', 'product__category').order_by('transaction_date')
        
        # Calculate totals
        total_revenue = 0
        total_cogs = 0
        
        # Prepare transaction data
        transaction_data = []
        for transaction in transactions:
            revenue = safe_float(transaction.quantity) * safe_float(transaction.unit_price)
            cost_per_unit = safe_float(transaction.product.purchase_price) if transaction.product else safe_float(transaction.unit_price * 0.7)
            cogs = safe_float(transaction.quantity) * cost_per_unit
            profit = revenue - cogs
                        
            total_revenue += revenue
            total_cogs += cogs
            
            # Format data with number translation
            formatted_data = [
                transaction.product.name if transaction.product else 'N/A',
                translate_numbers_to_arabic(transaction.product.item_id, language) if transaction.product else 'N/A',
                round(safe_float(transaction.quantity), 3),
                round(safe_float(transaction.unit_price), 3),
                round(cost_per_unit, 3),
                round(revenue, 3),
                round(cogs, 3),
                round(profit, 3),
                transaction.notes or ''
            ]
            
            transaction_data.append(formatted_data)
        
        total_profit = total_revenue - total_cogs
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"trip_{trip.trip_number}_{timestamp}"
        
        # Determine trip status with translation
        trip_status = labels['completed'] if getattr(trip, 'is_completed', False) else labels['pending']
        
        # Get vessel name in appropriate language
        vessel_name = get_vessel_name_by_language(trip.vessel, language)
        
        # Get transaction date
        local_dt = timezone.localtime(transaction.created_at).replace(tzinfo=None)
        
        # Translated metadata with number conversion
        metadata = {
            labels['export_date']: format_datetime_for_language(datetime.now(), language),
            labels['trip_number']: translate_numbers_to_arabic(str(trip.trip_number), language),
            labels['vessel']: vessel_name,
            labels['trip_date']: format_datetime_for_language(local_dt, language),
            labels['status']: trip_status,
            labels['passengers']: translate_numbers_to_arabic(str(safe_int(getattr(trip, 'passenger_count', 0))), language),
            labels['currency']: labels['jod'],
            labels['total_revenue_jod']: translate_numbers_to_arabic(f"{round(total_revenue, 3)}", language),
            labels['total_cogs_jod']: translate_numbers_to_arabic(f"{round(total_cogs, 3)}", language),
            labels['total_profit_jod']: translate_numbers_to_arabic(f"{round(total_profit, 3)}", language),
            # labels['profit_margin']: translate_numbers_to_arabic(f"{round((total_profit/total_revenue*100) if total_revenue > 0 else 0, 0)}%", language),
            labels['generated_by']: request.user.username
        }
        
        # Translated headers
        headers = [
            labels['product'], 
            labels['product_id'],
            labels['quantity'],
            labels['unit_selling_price_jod'],
            labels['cogs_per_unit_jod'],
            labels['revenue_jod'],
            labels['cogs_jod'],
            labels['profit_jod'],
            labels['notes']
        ]
        
        # Translated summary data with number conversion
        summary_data = {
            labels['total_items_sold']: translate_numbers_to_arabic(str(len(transaction_data)), language),
            labels['total_revenue_jod']: translate_numbers_to_arabic(f"{round(total_revenue, 3)}", language),
            labels['total_cogs_jod']: translate_numbers_to_arabic(f"{round(total_cogs, 3)}", language),
            labels['total_profit_jod']: translate_numbers_to_arabic(f"{round(total_profit, 3)}", language),
            # labels['profit_margin']: translate_numbers_to_arabic(f"{round((total_profit/total_revenue*100) if total_revenue > 0 else 0, 0)}%", language)
        }
        
        # Generate dynamic title
        trip_number_translated = translate_numbers_to_arabic(str(trip.trip_number), language)
        if language == 'ar':
            report_title = f"تقرير مبيعات رحلة {trip_number_translated}"
        else:
            report_title = f"Trip Sales {trip.trip_number} Report"
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title=report_title)
                
                # Add title with generation date
                generation_text = f"{labels['generated_on']} {format_datetime_for_language(datetime.now(), language)}"
                exporter.add_title(report_title, generation_text)
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(transaction_data)
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel single trip export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF - Simplified without CSS injection
            try:
                # Create context (no CSS injection needed)
                context = {
                    'title': report_title,
                    'metadata': metadata,
                    'tables': [{'title': f"{report_title} - {labels['transaction_details']}", 'id': 'trip_table', 'headers': headers, 'rows': transaction_data}],
                    'charts': [],
                    'summary_data': summary_data,
                    'orientation': 'landscape',
                    'language': language,  # This enables all RTL styling
                    'generation_date': format_datetime_for_language(datetime.now(), language),
                    'has_logo': False,
                    'generated_on_text': labels['generated_on'],
                    'report_info_text': labels['report_information'],
                    'summary_text': labels['summary'],
                    'company_logo_text': labels['company_logo'],
                    'no_data_text': labels['no_data_available'],
                }
                
                # Use template directly (no CSS injection)
                from django.template.loader import render_to_string
                import weasyprint
                import io
                
                template_name = 'frontend/exports/wide_report.html'
                html_string = render_to_string(template_name, context)
                
                html = weasyprint.HTML(string=html_string)
                
                buffer = io.BytesIO()
                html.write_pdf(target=buffer)
                buffer.seek(0)
                
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                response['Content-Length'] = len(buffer.getvalue())
                response.write(buffer.getvalue())
                
                return response
                
            except Exception as e:
                logger.error(f"PDF single trip export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Single trip export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

# ===============================================================================
# PURCHASE ORDER EXPORTS (LIST AND INDIVIDUAL)
# ===============================================================================

@login_required
@require_http_methods(["POST"])
def export_purchase_orders(request):
    """Export purchase orders list to Excel or PDF - Fixed with RTL support following individual export pattern"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get translated labels with language from request (for RTL support)
        labels = get_translated_labels(request, data)
        language = labels['language']
        
        # Get filters
        start_date, end_date = get_date_range_from_request(data)
        if not start_date or not end_date:
            return JsonResponse({'success': False, 'error': 'Invalid date range'})
            
        vessel_id = data.get('vessel_id')
        status = data.get('status')
        
        # Build query
        pos = PurchaseOrder.objects.select_related(
            'vessel', 'created_by'
        ).filter(
            po_date__gte=start_date,
            po_date__lte=end_date
        ).order_by('-po_date')
        
        # Apply filters
        if vessel_id:
            pos = pos.filter(vessel_id=vessel_id)
        if status:
            pos = pos.filter(status=status)
            
        # Prepare data
        table_data = []
        total_cost = 0
        total_items = 0
        
        for po in pos[:2000]:
            # Calculate PO total cost
            po_stats = Transaction.objects.filter(
                purchase_order=po,
                transaction_type='SUPPLY'
            ).aggregate(
                total_cost=Sum(F('quantity') * F('unit_price')),
                total_quantity=Sum('quantity'),
                transaction_count=Count('id')
            )
            
            cost = safe_float(po_stats['total_cost'])
            quantity = safe_int(po_stats['total_quantity'])
            transaction_count = safe_int(po_stats['transaction_count'])
            
            total_cost += cost
            total_items += quantity
            
            # Get vessel name in appropriate language
            vessel_name = get_vessel_name_by_language(po.vessel, language)
            
            # Determine PO status with translation
            po_status_key = 'completed' if getattr(po, 'is_completed', False) or getattr(po, 'status', '') == 'completed' else 'pending'
            po_status = labels[po_status_key]
            
            # Get current datetime for trip date metadata formatting
            local_dt = timezone.localtime(po.created_at).replace(tzinfo=None) if hasattr(po, 'created_at') else timezone.now().replace(tzinfo=None)
            
            formatted_data = [
                translate_numbers_to_arabic(str(po.po_number), language),
                vessel_name,
                format_date_for_language(local_dt, language),
                po_status,
                translate_numbers_to_arabic(str(quantity), language),
                translate_numbers_to_arabic(format_currency(cost, 3), language),
                translate_numbers_to_arabic(str(transaction_count), language),
                po.created_by.username if po.created_by else labels.get('system', 'System')
            ]
            table_data.append(formatted_data)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"purchase_orders_export_{timestamp}"
        
        # Get current datetime for metadata
        local_dt = timezone.localtime(timezone.now()).replace(tzinfo=None)
        
        # Metadata with translation
        metadata = {
            labels['export_date']: format_datetime_for_language(local_dt, language),
            labels['period']: f"{format_date_for_language(start_date, language)} - {format_date_for_language(end_date, language)}",
            labels['total_pos']: translate_numbers_to_arabic(str(len(table_data)), language),
            labels['total_cost_jod']: translate_numbers_to_arabic(format_currency(total_cost, 3), language),
            labels['currency']: labels['jod'],  # Apply new currency pattern
            labels['total_items_received']: translate_numbers_to_arabic(str(total_items), language),
            labels['generated_by']: request.user.username
        }
        
        # Headers with translation
        headers = [
            labels['po_number'],
            labels['vessel'],
            labels['po_date'],
            labels['status'],
            labels['total_quantity'],
            labels['total_cost_jod'],
            labels.get('transactions', 'Transactions'),
            labels['generated_by']
        ]
        
        # Summary data with translation
        summary_data = {
            labels['total_pos']: translate_numbers_to_arabic(str(len(table_data)), language),
            labels['total_cost_jod']: translate_numbers_to_arabic(format_currency(total_cost, 3), language),
            labels['total_items_received']: translate_numbers_to_arabic(str(total_items), language),
        }
        
        # Generate report title with translation
        if language == 'ar':
            report_title = labels['purchase_order_supply_report']
        else:
            report_title = labels['purchase_order_supply_report']
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title=report_title)
                
                generation_text = f"{labels['generated_on']} {format_datetime_for_language(datetime.now(), language)}"
                exporter.add_title(report_title, generation_text)
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(table_data)
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel PO export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF - Use the SAME pattern as working individual exports
            try:
                # Create context manually like the working individual exports
                context = {
                    'title': report_title,
                    'metadata': metadata,
                    'tables': [{'title': f"{report_title} - {labels.get('pos_overview', 'Purchase Orders Overview')}", 'id': 'pos_table', 'headers': headers, 'rows': table_data}],
                    'summary_data': summary_data,
                    'orientation': 'landscape',
                    'language': language,  # KEY: This enables RTL
                    'generation_date': format_datetime_for_language(datetime.now(), language),
                    'has_logo': False,
                    'generated_on_text': labels['generated_on'],
                    'report_info_text': labels['report_information'],
                    'summary_text': labels['summary'],
                    'company_logo_text': labels['company_logo'],
                    'no_data_text': labels['no_data_available'],
                }
                
                # Use the same pattern as working individual exports
                from django.template.loader import render_to_string
                import weasyprint
                import io
                
                template_name = 'frontend/exports/wide_report.html'
                html_string = render_to_string(template_name, context)
                
                html = weasyprint.HTML(string=html_string)
                
                buffer = io.BytesIO()
                html.write_pdf(target=buffer)
                buffer.seek(0)
                
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                response['Content-Length'] = len(buffer.getvalue())
                response.write(buffer.getvalue())
                
                return response
                
            except Exception as e:
                logger.error(f"PDF PO export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"PO export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

@login_required
@require_http_methods(["POST"])
def export_single_po(request, po_id):
    """Export individual purchase order details - Enhanced RTL support for metadata and summary sections"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get translated labels with language from request
        labels = get_translated_labels(request, data)
        language = labels['language']
        
        # Get PO
        po = get_object_or_404(PurchaseOrder, id=po_id)
        
        # Get PO transactions (supply transactions)
        transactions = Transaction.objects.filter(
            purchase_order=po,
            transaction_type='SUPPLY'
        ).select_related('product', 'product__category').order_by('transaction_date')
        
        # Calculate total cost
        total_cost = 0
        
        # Prepare transaction data
        transaction_data = []
        for transaction in transactions:
            cost = safe_float(transaction.quantity) * safe_float(transaction.unit_price)
            total_cost += cost
            
            # Format data with number translation
            formatted_data = [
                format_datetime_for_language(transaction.transaction_date, language),
                transaction.product.name if transaction.product else 'N/A',
                translate_numbers_to_arabic(transaction.product.item_id, language) if transaction.product else 'N/A',
                round(safe_float(transaction.quantity), 3),
                round(safe_float(transaction.unit_price), 3),
                round(cost, 3),
                transaction.notes or ''
            ]
            
            transaction_data.append(formatted_data)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"po_{po.po_number}_{timestamp}"
        
        # Determine PO status with translation
        po_status = labels['completed'] if getattr(po, 'is_completed', False) else labels['pending']
        
        # Get vessel name in appropriate language
        vessel_name = get_vessel_name_by_language(po.vessel, language)
        
        # Translated metadata with number conversion
        metadata = {
            labels['export_date']: format_datetime_for_language(datetime.now(), language),
            labels['po_number']: translate_numbers_to_arabic(str(po.po_number), language),
            labels['vessel']: vessel_name,
            labels['po_date']: format_date_for_language(po.po_date, language),
            labels['status']: po_status,
            labels['total_cost_jod']: translate_numbers_to_arabic(f"{total_cost:.3f}", language),
            labels['generated_by']: request.user.username
        }
        
        # Translated headers
        headers = [
            labels['time'],
            labels['product'],
            labels['product_id'], 
            labels['quantity'],
            labels['unit_cost'],
            labels['total_cost_jod'],
            labels['notes']
        ]
        
        # Translated summary with number conversion
        summary_data = {
            labels['total_items_received']: translate_numbers_to_arabic(str(len(transaction_data)), language),
            labels['total_cost_jod']: translate_numbers_to_arabic(f"{total_cost:.3f}", language),
            labels['average_cost_per_item_jod']: translate_numbers_to_arabic(f"{(total_cost / len(transaction_data)) if transaction_data else 0:.3f}", language)
        }
        
        # Generate dynamic title
        po_number_translated = translate_numbers_to_arabic(str(po.po_number), language)
        if language == 'ar':
            report_title = f"تقرير أمر شراء {po_number_translated}"
        else:
            report_title = f"PO {po.po_number} Report"
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title=report_title)
                
                # Add title with generation date
                generation_text = f"{labels['generated_on']} {format_datetime_for_language(datetime.now(), language)}"
                exporter.add_title(report_title, generation_text)
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(transaction_data)
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel single PO export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF - Simplified without CSS injection
            try:
                # Create context (no CSS injection needed)
                context = {
                    'title': report_title,
                    'metadata': metadata,
                    'tables': [{'title': f"{report_title} - {labels['item_details']}", 'id': 'po_table', 'headers': headers, 'rows': transaction_data}],
                    'charts': [],
                    'summary_data': summary_data,
                    'orientation': 'portrait',
                    'language': language,  # This enables all RTL styling
                    'generation_date': format_datetime_for_language(datetime.now(), language),
                    'has_logo': False,
                    'generated_on_text': labels['generated_on'],
                    'report_info_text': labels['report_information'],
                    'summary_text': labels['summary'],
                    'company_logo_text': labels['company_logo'],
                    'no_data_text': labels['no_data_available'],
                }
                
                # Use template directly (no CSS injection)
                from django.template.loader import render_to_string
                import weasyprint
                import io
                
                template_name = 'frontend/exports/standard_report.html'
                html_string = render_to_string(template_name, context)
                
                html = weasyprint.HTML(string=html_string)
                
                buffer = io.BytesIO()
                html.write_pdf(target=buffer)
                buffer.seek(0)
                
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                response['Content-Length'] = len(buffer.getvalue())
                response.write(buffer.getvalue())
                
                return response
                
            except Exception as e:
                logger.error(f"PDF single PO export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Single PO export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

# ===============================================================================
# MONTHLY REPORT EXPORTS
# ===============================================================================

@login_required
@require_http_methods(["POST"])
def export_monthly_report(request):
    """Export monthly performance report with transfer tracking"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get month and year
        selected_month = safe_int(data.get('month', datetime.now().month))
        selected_year = safe_int(data.get('year', datetime.now().year))
        
        if not (1 <= selected_month <= 12) or selected_year < 2020:
            return JsonResponse({'success': False, 'error': 'Invalid month or year'})
        
        # Calculate date range for the month
        start_date = datetime(selected_year, selected_month, 1).date()
        if selected_month == 12:
            end_date = datetime(selected_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(selected_year, selected_month + 1, 1).date() - timedelta(days=1)
        
        # Get month name
        month_name = calendar.month_name[selected_month]
        
        # Calculate performance for each vessel
        vessels = Vessel.objects.filter(active=True)
        vessel_performance = []
        monthly_revenue = 0
        monthly_costs = 0
        monthly_transfer_out = 0
        monthly_transfer_in = 0
        
        for vessel in vessels:
            # Get sales (revenue) for the month
            sales = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='SALE',
                transaction_date__gte=start_date,
                transaction_date__lte=end_date
            ).aggregate(
                revenue=Sum(F('quantity') * F('unit_price')),
                count=Count('id')
            )
            
            # Get supplies (costs) for the month
            supplies = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='SUPPLY',
                transaction_date__gte=start_date,
                transaction_date__lte=end_date
            ).aggregate(
                costs=Sum(F('quantity') * F('unit_price')),
                count=Count('id')
            )
            
            # Get transfer out for the month
            transfers_out = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='TRANSFER_OUT',
                transaction_date__gte=start_date,
                transaction_date__lte=end_date
            )
            transfer_out_total = sum(calculate_transfer_amounts(t) for t in transfers_out)
            
            # Get transfer in for the month
            transfers_in = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='TRANSFER_IN',
                transaction_date__gte=start_date,
                transaction_date__lte=end_date
            )
            transfer_in_total = sum(calculate_transfer_amounts(t) for t in transfers_in)
            
            revenue = safe_float(sales['revenue'])
            costs = safe_float(supplies['costs'])
            profit = revenue - costs
            profit_margin = (profit / revenue * 100) if revenue > 0 else 0
            
            monthly_revenue += revenue
            monthly_costs += costs
            monthly_transfer_out += transfer_out_total
            monthly_transfer_in += transfer_in_total
            
            vessel_performance.append([
                vessel.name,
                format_currency(revenue, 3),
                format_currency(costs, 3),
                format_currency(profit, 3),
                format_percentage(profit_margin),
                safe_int(sales['count']),
                safe_int(supplies['count']),
                format_currency(transfer_out_total, 3),
                format_currency(transfer_in_total, 3)
            ])
        
        monthly_profit = monthly_revenue - monthly_costs
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"monthly_report_{selected_year}_{selected_month:02d}_{timestamp}"
        
        # Metadata
        metadata = {
            'Report Period': f"{month_name} {selected_year}",
            'Export Date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'Date Range': f"{format_date(start_date)} to {format_date(end_date)}",
            'Total Revenue (JOD)': format_currency(monthly_revenue, 3),
            'Total Costs (JOD)': format_currency(monthly_costs, 3),
            'Total Profit (JOD)': format_currency(monthly_profit, 3),
            'Total Transfer Out (JOD)': format_currency(monthly_transfer_out, 3),
            'Total Transfer In (JOD)': format_currency(monthly_transfer_in, 3),
            'Overall Profit Margin': format_percentage((monthly_profit / monthly_revenue * 100) if monthly_revenue > 0 else 0),
            'Generated By': request.user.username
        }
        
        headers = [
            'Vessel', 'Revenue (JOD)', 'Costs (JOD)', 'Profit (JOD)', 
            'Profit Margin (%)', 'Sales Count', 'Supply Count',
            'Transfer Out (JOD)', 'Transfer In (JOD)'
        ]
        
        # Create summary data 
        summary_data = {
            'Total Vessels': len(vessel_performance),
            'Total Revenue (JOD)': format_currency(monthly_revenue, 3),
            'Total Profit (JOD)': format_currency(monthly_profit, 3),
            'Overall Profit Margin': format_percentage((monthly_profit / monthly_revenue * 100) if monthly_revenue > 0 else 0),
            'Net Transfers': format_currency(monthly_transfer_out - monthly_transfer_in, 3)
        }
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title="Monthly Report")
                exporter.add_title(f"Monthly Report - {month_name} {selected_year}", f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(vessel_performance)
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel monthly report export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF
            try:
                exporter = create_weasy_exporter_for_data("Monthly Report", "normal")
                exporter.add_metadata(metadata)
                exporter.add_table(headers, vessel_performance, table_title=f"Monthly Performance - {month_name} {selected_year}")
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.pdf")
                
            except Exception as e:
                logger.error(f"PDF monthly report export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Monthly report export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

# ===============================================================================
# DAILY REPORT EXPORTS
# ===============================================================================

@login_required
@require_http_methods(["POST"])
def export_daily_report(request):
    """Export daily performance report"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get date
        selected_date = data.get('date')
        if selected_date:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        else:
            selected_date = timezone.now().date()
        
        # Calculate performance for each vessel for the selected day
        vessels = Vessel.objects.filter(active=True)
        vessel_performance = []
        daily_revenue = 0
        daily_costs = 0
        
        for vessel in vessels:
            # Get sales (revenue) for the day
            sales = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='SALE',
                transaction_date=selected_date
            ).aggregate(
                revenue=Sum(F('quantity') * F('unit_price')),
                count=Count('id')
            )
            
            # Get supplies (costs) for the day
            supplies = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='SUPPLY',
                transaction_date=selected_date
            ).aggregate(
                costs=Sum(F('quantity') * F('unit_price')),
                count=Count('id')
            )
            
            revenue = safe_float(sales['revenue'])
            costs = safe_float(supplies['costs'])
            profit = revenue - costs
            profit_margin = (profit / revenue * 100) if revenue > 0 else 0
            
            # Only include vessels with activity
            if revenue > 0 or costs > 0:
                daily_revenue += revenue
                daily_costs += costs
                
                vessel_performance.append([
                    vessel.name,
                    format_currency(revenue, 3),
                    format_currency(costs, 3),
                    format_currency(profit, 3),
                    format_percentage(profit_margin),
                    safe_int(sales['count']),
                    safe_int(supplies['count'])
                ])
        
        daily_profit = daily_revenue - daily_costs
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"daily_report_{selected_date.strftime('%Y%m%d')}_{timestamp}"
        
        # Metadata
        metadata = {
            'Report Date': format_date(selected_date),
            'Export Date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'Total Revenue (JOD)': format_currency(daily_revenue, 3),
            'Total Costs (JOD)': format_currency(daily_costs, 3),
            'Total Profit (JOD)': format_currency(daily_profit, 3),
            'Overall Profit Margin': format_percentage((daily_profit / daily_revenue * 100) if daily_revenue > 0 else 0),
            'Generated By': request.user.username
        }
        
        headers = [
            'Vessel', 'Revenue (JOD)', 'Costs (JOD)', 'Profit (JOD)', 
            'Profit Margin (%)', 'Sales Count', 'Supply Count'
        ]
        
        # Create summary data
        summary_data = {
            'Active Vessels': len(vessel_performance),
            'Total Revenue (JOD)': format_currency(daily_revenue, 3),
            'Total Profit (JOD)': format_currency(daily_profit, 3),
            'Overall Profit Margin': format_percentage((daily_profit / daily_revenue * 100) if daily_revenue > 0 else 0)
        }
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title="Daily Report")
                exporter.add_title(f"Daily Report - {format_date(selected_date)}", f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(vessel_performance)
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel daily report export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF
            try:
                exporter = create_weasy_exporter_for_data("Daily Report", "normal")
                exporter.add_metadata(metadata)
                exporter.add_table(headers, vessel_performance, table_title=f"Daily Performance - {format_date(selected_date)}")
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.pdf")
                
            except Exception as e:
                logger.error(f"PDF daily report export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Daily report export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})

# ===============================================================================
# ANALYTICS EXPORT
# ===============================================================================

@login_required
@require_http_methods(["POST"])
def export_analytics(request):
    """Export analytics report with performance metrics and charts"""
    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'excel')
        
        # Get date range (default to last 30 days)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
        
        # Override with request data if provided
        if data.get('start_date'):
            start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%d').date()
        if data.get('end_date'):
            end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%d').date()
        
        # Calculate comprehensive analytics for each vessel
        vessels = Vessel.objects.filter(active=True)
        vessel_analytics = []
        total_revenue = 0
        total_costs = 0
        total_pos = 0
        po_analytics = []
        
        for vessel in vessels:
            # Get sales data
            sales_stats = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='SALE',
                transaction_date__gte=start_date,
                transaction_date__lte=end_date
            ).aggregate(
                revenue=Sum(F('quantity') * F('unit_price')),
                sales_count=Count('id')
            )
            
            # Get supply data
            supply_stats = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='SUPPLY',
                transaction_date__gte=start_date,
                transaction_date__lte=end_date
            ).aggregate(
                costs=Sum(F('quantity') * F('unit_price')),
                supply_count=Count('id')
            )
            
            # Get PO data
            po_stats = PurchaseOrder.objects.filter(
                vessel=vessel,
                po_date__gte=start_date,
                po_date__lte=end_date
            ).aggregate(
                po_count=Count('id')
            )
            
            # Calculate PO costs from related transactions
            po_cost = Transaction.objects.filter(
                vessel=vessel,
                transaction_type='SUPPLY',
                transaction_date__gte=start_date,
                transaction_date__lte=end_date,
                purchase_order__isnull=False
            ).aggregate(
                total=Sum(F('quantity') * F('unit_price'))
            )['total'] or 0
            
            revenue = safe_float(sales_stats['revenue'])
            costs = safe_float(supply_stats['costs'])
            profit_margin = ((revenue - costs) / revenue * 100) if revenue > 0 else 0
            avg_transaction = revenue / safe_int(sales_stats['sales_count']) if safe_int(sales_stats['sales_count']) > 0 else 0
            
            total_revenue += revenue
            total_costs += costs
            total_pos += safe_int(po_stats['po_count'])
            
            # Only include vessels with activity
            if revenue > 0 or costs > 0 or po_cost > 0:
                vessel_analytics.append([
                    vessel.name,
                    revenue,  # Store as number, not formatted string
                    costs,    # Store as number, not formatted string
                    profit_margin,  # Store as number, not formatted string
                    safe_int(sales_stats['sales_count']),
                    avg_transaction,  # Store as number
                    safe_int(po_stats['po_count']),
                    po_cost   # Store as number
                ])
                
                # For PO chart data
                if po_cost > 0:
                    po_analytics.append((vessel.name[:15], po_cost))
        
        # Sort by revenue (descending)
        vessel_analytics.sort(key=lambda x: float(x[1]), reverse=True)
        
        total_profit = total_revenue - total_costs
        
        # Calculate additional statistics
        total_stats = Transaction.objects.filter(
            transaction_date__gte=start_date,
            transaction_date__lte=end_date
        ).aggregate(
            transaction_count=Count('id')
        )
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"analytics_report_{timestamp}"
        
        # Metadata
        metadata = {
            'Analysis Period': f"{format_date(start_date)} to {format_date(end_date)}",
            'Total Revenue (JOD)': format_currency(total_revenue, 3),
            'Total Costs (JOD)': format_currency(total_costs, 3),
            'Total Profit (JOD)': format_currency(total_profit, 3),
            'Total Transactions': safe_int(total_stats['transaction_count']),
            'Total Purchase Orders': total_pos,
            'Average Daily Revenue (JOD)': format_currency((total_revenue / (end_date - start_date).days), 3),
            'Overall Profit Margin': format_percentage((total_profit / total_revenue * 100) if total_revenue > 0 else 0),
            'Generated By': request.user.username
        }
        
        headers = [
            'Vessel', 'Revenue (JOD)', 'Costs (JOD)', 'Profit Margin (%)', 
            'Sales Count', 'Avg Transaction Value (JOD)', 'PO Count', 'PO Total Cost (JOD)'
        ]
        
        # Create summary data
        summary_data = {
            'Analysis Period': f"{(end_date - start_date).days} days",
            'Total Revenue (JOD)': format_currency(total_revenue, 3),
            'Overall Profit Margin': format_percentage((total_profit / total_revenue * 100) if total_revenue > 0 else 0),
            'Best Performing Vessel': vessel_analytics[0][0] if vessel_analytics else None,
            'Total Purchase Orders': total_pos
        }
        
        if export_format == 'excel':
            try:
                exporter = ExcelExporter(title="Analytics Report")
                exporter.add_title("Analytics Report", f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                exporter.add_metadata(metadata)
                exporter.add_headers(headers)
                exporter.add_data_rows(vessel_analytics)  # Now contains raw numbers
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.xlsx")
                
            except Exception as e:
                logger.error(f"Excel analytics export error: {e}")
                return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})
        
        else:  # PDF - Now landscape with charts
            try:
                exporter = create_weasy_exporter(title="Analytics Report", template_type="analytics", orientation="landscape")
                exporter.add_metadata(metadata)
                
                # Add charts for analytics
                if vessel_analytics:
                    try:
                        # Revenue chart - use raw numbers
                        revenue_data = [(row[0][:15], float(row[1])) for row in vessel_analytics[:8]]
                        if revenue_data:
                            exporter.add_chart(revenue_data, 'bar', 'Revenue by Vessel (JOD)', 'revenue_chart')
                        
                        # Profit margin chart - use raw numbers  
                        profit_data = [(row[0][:15], float(row[3])) for row in vessel_analytics[:8]]
                        if profit_data:
                            exporter.add_chart(profit_data, 'bar', 'Profit Margin by Vessel (%)', 'profit_chart')
                        
                        # Add PO chart
                        if po_analytics:
                            # Sort PO data by cost (descending) and take top 8
                            po_analytics.sort(key=lambda x: x[1], reverse=True)
                            po_chart_data = po_analytics[:8]
                            exporter.add_chart(po_chart_data, 'bar', 'Purchase Order Costs by Vessel (JOD)', 'po_chart')
                            
                    except Exception as chart_error:
                        logger.warning(f"Could not create charts for analytics: {chart_error}")
                
                # Format data for PDF display
                formatted_vessel_analytics = []
                for row in vessel_analytics:
                    formatted_row = [
                        row[0],  # Vessel name
                        format_currency(row[1], 3),  # Revenue
                        format_currency(row[2], 3),  # Costs
                        format_percentage(row[3]),   # Profit margin
                        row[4],  # Sales count
                        format_currency(row[5], 3),  # Avg transaction
                        row[6],  # PO count
                        format_currency(row[7], 3)   # PO cost
                    ]
                    formatted_vessel_analytics.append(formatted_row)
                
                exporter.add_table(headers, formatted_vessel_analytics, table_title="Vessel Performance Analytics")
                exporter.add_summary(summary_data)
                
                return exporter.get_response(f"{filename_base}.pdf")
                
            except Exception as e:
                logger.error(f"PDF analytics export error: {e}")
                return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})
                
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Analytics export error: {e}")
        return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})